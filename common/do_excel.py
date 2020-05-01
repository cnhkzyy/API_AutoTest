from openpyxl import load_workbook
from common.replace_variable import *
import json, re


class DoExcel:

    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = load_workbook(self.file_path)
        self.sh = self.wb["case_datas"]


    #读取所有测试数据
    def read_all_caseData(self):
        #得到最大行、最大列
        max_row = self.sh.max_row
        max_column = self.sh.max_column
        #得到合并单元格的范围
        m_rows = []
        m_cols = []
        for m_area in self.sh.merged_cells:
            r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
            m_rows.append((r1, r2))
            m_cols.extend([c1, c2])
        m_rows = set(m_rows)
        m_cols = set(m_cols)

        #获取所有的测试数据，将其以字典的形式存储在列表中
        all_case_data = []
        title = [self.sh.cell(1, column).value for column in range(1, max_column + 1)]
        #获取合并行数据
        for m_row in m_rows:
            case_data = []
            for column in range(1, max_column + 1):
                for row in range(m_row[0], m_row[1] + 1):
                    if row == m_row[0]:
                        value = [[self.sh.cell(row, column).value]]
                    else:
                        for item in value:
                            item.append(self.sh.cell(row, column).value)
                case_data.extend(value)
            all_case_data.append(dict(zip(title, case_data)))

        #最大行范围和合并范围求差集，就是非合并行
        max_row_range = set(range(2, max_row + 1))
        for m_row in m_rows:
            m_row_range = set(range(m_row[0], m_row[1] + 1))
            max_row_range = max_row_range.difference(m_row_range)

        #获取非合并行数据
        for row in max_row_range:
            case_data = []
            for column in range(1, max_column + 1):
                value = [self.sh.cell(row, column).value]
                case_data.append(value)
            all_case_data.append(dict(zip(title, case_data)))

        #排序：根据case_id的数字从大到小
        all_case_data = sorted(all_case_data, key=lambda case_data: case_data["case_id"][0])

        #对检查的字符串做字典的转化，最终以字典的形式存储在列表中
        for case_data in all_case_data:
            self._check_convert_dict(case_data)
        print(all_case_data)
        return all_case_data


    #将检查项转化为字典
    def _check_convert_dict(self, case_data):
        for i in range(len(case_data["case_id"])):
            check_list = case_data["check"]
            new_check_list = []
            for check in check_list:
                new_check = {}
                items = re.findall("(\$.*?):.*?\"(.*?)\"", check)
                for item in items:
                    key, value = item
                    new_check[key] = value
                new_check_list.append(new_check)
        case_data["check"] = new_check_list




if __name__ == "__main__":
    all_case_data = DoExcel(r"E:\virtual_workshop\Python_API\test_datas\api_info.xlsx").read_all_caseData()
